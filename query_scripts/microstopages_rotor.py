from openpyxl import load_workbook
import os
from datetime import datetime
import math
import time

DATEFORMAT = "%m/%d/%Y %H:%M:%S %p"
DICTIONARY = {}

def count_time(time1, time2):
    """counts minutes between two times"""
    t1 = datetime.strptime(time1, DATEFORMAT)
    t2 = datetime.strptime(time2, DATEFORMAT)
    secs = (abs(t2 - t1).seconds)
    return secs

def fill_dict(zipped_rows, unique_equipment):
    """creates dictionary by connecting unique equipment values with rows and get rid of equipment in tuples"""
    for i in zipped_rows:
        for eq in unique_equipment:
            if i[0] == eq:
                if eq not in DICTIONARY:
                    DICTIONARY[eq] = list()
                DICTIONARY[eq].append(i[1:])
    return DICTIONARY

def plc_error(text,type):
        if f"<{type}>" in text:
            splitted = text.split(f"{type}")
            extract = splitted[1][1:-2]
            return extract
        else:
            return "NULL"
#NAMING FOR THE NEW FILE
timestr = time.strftime("%d-%m-%Y_%H-%M-POSTOJE-ROTOR")
file_name = timestr

def main():
    #NAME OF THE OLD FILE
    file_name = '2.xlsx'
    full_file = os.path.abspath(os.path.join('', file_name))
    wb = load_workbook(full_file)
    ws = wb.active

    ws["G1"] = "DESCRIPTION"
    ws["H1"] = "CODE"
    ws["I1"] = "STATE"
    ws["J1"] = "DATE"
    ws["K1"] = "TERMINAL"
    ws["L1"] = "LINE"
    ws["M1"] = "SERIAL"
    ws["N1"] = "PRODUCT"
    ws["O1"] = "TIME"  
            
    """LISTS"""
    state_list = []
    eq_list = []
    row_list = []
    event_list = []
    unique_eq = set()

    for number in range(2, ws.max_row):
        test_string = (ws["A" + str(number)].value)
        equipment = plc_error(test_string, "Equipment")
        desc = plc_error(test_string, "PLCErrorDesc")
        code = plc_error(test_string, "PLCErrorCode")
        state = plc_error(test_string, "State")
        eventdatetime = plc_error(test_string, "EventDateTime")
        workcenter = plc_error(test_string, "WorkCenter")
        unique_eq.add(equipment)

        if state != "NULL":
            state_list.append(int(state))
            row_list.append(number)
            eq_list.append(equipment)
            event_list.append(eventdatetime)
#            zipped = list(zip(eq_list, state_list, event_list, row_list))
            zipped = zip(eq_list, state_list, event_list, row_list)

        ws.cell(row=number, column=7, value=desc)
        ws.cell(row=number, column=8, value=code)
        ws.cell(row=number, column=9, value=state)
        ws.cell(row=number, column=10, value=eventdatetime)
        ws.cell(row=number, column=11, value=equipment)
        ws.cell(row=number, column=12, value=workcenter)

    zipped = list(zipped)

    new_dict = fill_dict(zipped, unique_eq)

    for _,val in new_dict.items():
        for i in range(len(val) - 1):
            if val[i][0] == 1:
                if val[i + 1][0] == 0:
                    time = count_time(val[i][1], val[i + 1][1])
                    count_min = math.floor(time / 60)
                    count_sec = time % 60
                    if count_sec < 10:
                        count_sec = f"0{count_sec}"
                    new_time = f"{count_min}:{count_sec}"
                    ws.cell(row=val[i][2], column=15, value=new_time) # TIME



    wb.save(file_name+".xlsx")
 
    
if __name__ == "__main__":
    main()
